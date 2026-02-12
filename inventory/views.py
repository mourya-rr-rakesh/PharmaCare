from django.shortcuts import render

# Create your views here.


import json
from datetime import date, timedelta
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from .models import Medicine
import openpyxl
from openpyxl.styles import PatternFill, Font
from io import BytesIO
from django.http import HttpResponse

def auth_required(request):
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Unauthorized'}, status=401)
    return None


@csrf_exempt
def medicines(request):
    auth = auth_required(request)
    if auth:
        return auth

    user = request.user

    if request.method == 'GET':
        meds = Medicine.objects.filter(user=user)
        data = []
        for m in meds:
            data.append({
                '_id': m.id,
                'name': m.name,
                'type': m.type,
                'composition': m.composition,
                'mfgDate': m.mfg_date,
                'expDate': m.exp_date,
                'mrp': m.mrp,
                'buyPrice': m.buy_price,
                'sellPrice': m.sell_price,
                'stock': m.stock
            })
        return JsonResponse({'medicines': data})

    if request.method == 'POST':
        data = json.loads(request.body)

        # Check for duplicate medicine with same name, composition, and expiry date
        # Allow duplicate if expiry date is different
        existing_med = Medicine.objects.filter(
            user=user,
            name=data['name'],
            composition=data.get('composition', ''),
            exp_date=data['expDate']
        ).first()
        
        if existing_med:
            return JsonResponse({
                'error': f"Medicine '{data['name']}' with expiry date {data['expDate']} already exists! Different expiry date allowed."
            }, status=400)

        med = Medicine.objects.create(
            user=user,
            name=data['name'],
            type=data.get('type', ''),
            composition=data.get('composition', ''),
            mfg_date=data.get('mfgDate') or None,
            exp_date=data['expDate'],
            mrp=data['mrp'],
            buy_price=data['buyPrice'],
            sell_price=data['sellPrice'],
            stock=data['stock']
        )
        return JsonResponse({'message': 'Medicine added successfully'})

    return JsonResponse({'error': 'Method not allowed'}, status=405)


@csrf_exempt
def medicine_detail(request, id):
    auth = auth_required(request)
    if auth:
        return auth

    try:
        med = Medicine.objects.get(id=id, user=request.user)
    except Medicine.DoesNotExist:
        return JsonResponse({'error': 'Medicine not found'}, status=404)

    if request.method == 'PUT':
        data = json.loads(request.body)
        
        # Check for duplicate medicine with same name, composition, and expiry date (excluding current medicine)
        existing_med = Medicine.objects.filter(
            user=request.user,
            name=data['name'],
            composition=data.get('composition', ''),
            exp_date=data['expDate']
        ).exclude(id=id).first()
        
        if existing_med:
            return JsonResponse({
                'error': f"Medicine '{data['name']}' with expiry date {data['expDate']} already exists! Different expiry date allowed."
            }, status=400)
        
        med.name = data['name']
        med.type = data.get('type', '')
        med.composition = data.get('composition', '')
        med.mfg_date = data.get('mfgDate') or None
        med.exp_date = data['expDate']
        med.mrp = data['mrp']
        med.buy_price = data['buyPrice']
        med.sell_price = data['sellPrice']
        med.stock = data['stock']
        med.save()
        return JsonResponse({'message': 'Medicine updated successfully'})

    if request.method == 'DELETE':
        med.delete()
        return JsonResponse({'message': 'Medicine deleted successfully'})

    return JsonResponse({'error': 'Method not allowed'}, status=405)



# Soon-Expiring medicines APIs

@csrf_exempt
def soon_expiring(request):
    auth = auth_required(request)
    if auth:
        return auth

    today = date.today()
    limit = today + timedelta(days=150)  # ~5 months

    meds = Medicine.objects.filter(
        user=request.user,
        exp_date__lte=limit
    )

    data = []
    for m in meds:
        data.append({
            '_id': m.id,
            'name': m.name,
            'type': m.type,
            'composition': m.composition,
            'mfgDate': m.mfg_date,
            'expDate': m.exp_date,
            'mrp': m.mrp,
            'buyPrice': m.buy_price,
            'sellPrice': m.sell_price,
            'stock': m.stock
        })

    return JsonResponse({'medicines': data})


@csrf_exempt
def export_medicines(request):
    """Export medicines to Excel file"""
    auth = auth_required(request)
    if auth:
        return auth
    
    try:
        meds = Medicine.objects.filter(user=request.user)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Medicines"
        
        # Add headers
        headers = ["ID", "Medicine Name", "Type", "Composition", "MFG Date", "EXP Date", "MRP", "Buy Price", "Sell Price", "Stock"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Add data rows
        for row, m in enumerate(meds, 2):
            ws.cell(row=row, column=1, value=m.id)
            ws.cell(row=row, column=2, value=m.name)
            ws.cell(row=row, column=3, value=m.type)
            ws.cell(row=row, column=4, value=m.composition)
            ws.cell(row=row, column=5, value=m.mfg_date.strftime('%Y-%m-%d') if m.mfg_date else '')
            ws.cell(row=row, column=6, value=m.exp_date.strftime('%Y-%m-%d') if m.exp_date else '')
            ws.cell(row=row, column=7, value=m.mrp)
            ws.cell(row=row, column=8, value=m.buy_price)
            ws.cell(row=row, column=9, value=m.sell_price)
            ws.cell(row=row, column=10, value=m.stock)
        
        # Adjust column widths
        for col in range(1, 11):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
        
        # Write to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="medicines_{date.today()}.xlsx"'
        return response
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def import_medicines(request):
    """Bulk import medicines"""
    auth = auth_required(request)
    if auth:
        return auth
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        medicines_data = data.get('medicines', [])
        
        if not medicines_data:
            return JsonResponse({'error': 'No medicines provided'}, status=400)
        
        created_count = 0
        for med_data in medicines_data:
            try:
                mfg_date_val = med_data.get('mfgDate')
                exp_date_val = med_data.get('expDate')
                mrp_val = med_data.get('mrp', 0)
                buy_price_val = med_data.get('buyPrice', 0)
                sell_price_val = med_data.get('sellPrice', 0)
                stock_val = med_data.get('stock', 0)
                
                # Convert to proper types
                try:
                    mrp_val = float(mrp_val) if mrp_val else 0
                except:
                    mrp_val = 0
                
                try:
                    buy_price_val = float(buy_price_val) if buy_price_val else 0
                except:
                    buy_price_val = 0
                
                try:
                    sell_price_val = float(sell_price_val) if sell_price_val else 0
                except:
                    sell_price_val = 0
                
                try:
                    stock_val = int(stock_val) if stock_val else 0
                except:
                    stock_val = 0
                
                Medicine.objects.create(
                    user=request.user,
                    name=med_data.get('name', ''),
                    type=med_data.get('type', ''),
                    composition=med_data.get('composition', ''),
                    mfg_date=mfg_date_val if mfg_date_val else None,
                    exp_date=exp_date_val if exp_date_val else None,
                    mrp=mrp_val,
                    buy_price=buy_price_val,
                    sell_price=sell_price_val,
                    stock=stock_val
                )
                created_count += 1
            except Exception as e:
                print(f"Error creating medicine: {e}")
                continue
        
        return JsonResponse({
            'message': f'Successfully imported {created_count} medicines',
            'count': created_count
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def search_medicines(request):
    """Search medicines by name or composition"""
    auth = auth_required(request)
    if auth:
        return auth
    
    query = request.GET.get('q', '').strip()
    if not query:
        return JsonResponse({'medicines': []})
    
    try:
        from django.db.models import Q
        query_filter = Q(name__icontains=query) | Q(composition__icontains=query)
        meds = Medicine.objects.filter(
            query_filter,
            user=request.user
        )
        
        data = []
        for m in meds:
            data.append({
                '_id': m.id,
                'name': m.name,
                'type': m.type,
                'composition': m.composition,
                'mfgDate': m.mfg_date,
                'expDate': m.exp_date,
                'mrp': m.mrp,
                'buyPrice': m.buy_price,
                'sellPrice': m.sell_price,
                'stock': m.stock
            })
        
        return JsonResponse({'medicines': data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

from django.shortcuts import render, get_object_or_404
from .models import Medicine

def dashboard(request):
    medicine_id = request.GET.get('medicine_id')
    medicine = None

    if medicine_id:
        medicine = get_object_or_404(Medicine, id=medicine_id)

    return render(request, 'dashboard.html', {
        'medicine': medicine
    })